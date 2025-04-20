from django.contrib import admin
from .models import Profile, Cart, CartItems, Order, OrderItem, Product
import openpyxl
from django.http import HttpResponse
from django.urls import path
from django.utils.html import format_html
from django.utils.timezone import now, timedelta
from openpyxl.utils import get_column_letter


# Register your models here.

class ProfileAdmin(admin.ModelAdmin):
    list_display = ("user", "formatted_email_verified", "phone", "formatted_address", "total_orders", "total_spent")

    def formatted_email_verified(self, obj):
        return "✅ Yes" if obj.is_email_verified else "❌ No"
    formatted_email_verified.short_description = "Email Verified"

    def formatted_address(self, obj):
        return f"{obj.address}, {obj.city}, {obj.state}, {obj.country}" if obj.address else "N/A"
    formatted_address.short_description = "Address"

    def total_orders(self, obj):
        return obj.total_orders()
    total_orders.short_description = "Total Orders"

    def total_spent(self, obj):
        return obj.total_spent()
    total_spent.short_description = "Total Spent (₹)"

admin.site.register(Profile, ProfileAdmin)
    
class CartAdmin(admin.ModelAdmin):
    list_display = ("user", "formatted_total_price", "formatted_coupon", "formatted_is_paid", "razor_pay_order_id")

    def formatted_total_price(self, obj):
        return f"₹{obj.get_cart_total()}"
    formatted_total_price.short_description = "Total Price"

    def formatted_coupon(self, obj):
        return obj.coupon.coupon_code if obj.coupon else "No Coupon"
    formatted_coupon.short_description = "Coupon"

    def formatted_is_paid(self, obj):
        return "✅ Yes" if obj.is_paid else "❌ No"
    formatted_is_paid.short_description = "Paid"

admin.site.register(Cart, CartAdmin)

class CartItemsAdmin(admin.ModelAdmin):
    list_display = ("cart", "formatted_product", "formatted_color", "formatted_size", "quantity", "formatted_total_price")

    def formatted_product(self, obj):
        return obj.product.product_name if obj.product else "Product Deleted"
    formatted_product.short_description = "Product"

    def formatted_color(self, obj):
        return obj.color_variant.color_name if obj.color_variant else "No Color"
    formatted_color.short_description = "Color"

    def formatted_size(self, obj):
        return obj.size_variant.size_name if obj.size_variant else "No Size"
    formatted_size.short_description = "Size"

    def formatted_total_price(self, obj):
        return f"₹{obj.get_product_price()}"
    formatted_total_price.short_description = "Total Price"

admin.site.register(CartItems, CartItemsAdmin)

class OrderAdmin(admin.ModelAdmin):
    list_display = ("uid", "user", "formatted_total_price", "status", "coupon", "formatted_date")
    ordering = ("-created_at",) 
    change_list_template = "admin/orders_changelist.html"  # Custom template to inject the button

    def full_name(self, obj):
        return f"{obj.user.first_name} {obj.user.last_name}".strip()
    full_name.short_description = "Customer Name"

    def formatted_total_price(self, obj):
        return f"₹{obj.total_price:.2f}"  # Formats price
    formatted_total_price.short_description = "Total Price"

    def formatted_date(self, obj):
        return obj.created_at.strftime("%d-%m-%Y %I:%M %p")  # Shows Date & Time
    formatted_date.short_description = "Order Date & Time"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('download-report/', self.admin_site.admin_view(self.download_excel), name='download_order_report'),
        ]
        return custom_urls + urls

    def download_excel(self, request):
        # Filter by date range
        date_range = request.GET.get('range', 'all')
        orders = Order.objects.select_related('user').order_by('-created_at')

        if date_range == '7days':
            orders = orders.filter(created_at__gte=now() - timedelta(days=7))
        elif date_range == '30days':
            orders = orders.filter(created_at__gte=now() - timedelta(days=30))

        # Create Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Order Report"

        # Write headers for the Excel file
        headers = [
        "Order ID",
        "User",
        "Email",
        "Phone Number",
        "Total Price",
        "Order Status",
        "Coupon Code",
        "Order Date",
        "Razorpay Order ID",
        "Razorpay Payment ID",
        ]
        ws.append(headers)

        # Data rows
        for order in orders:
            full_name = f"{order.user.first_name} {order.user.last_name}"
            ws.append([
                str(order.uid),
                full_name,
                order.user.email,
                order.user.profile.phone if hasattr(order.user, 'profile') else '',
                float(order.total_price),
                order.status,
                order.coupon.coupon_code if order.coupon else '',
                order.created_at.strftime("%Y-%m-%d %H:%M"),
                order.razor_pay_order_id if hasattr(order, 'razor_pay_order_id') else 'N/A',
                order.razor_pay_payment_id if hasattr(order, 'razor_pay_payment_id') else 'N/A'
            ])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Set the response for the file download
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=order_report.xlsx'
        wb.save(response)
        return response

admin.site.register(Order, OrderAdmin)


class OrderItemAdmin(admin.ModelAdmin):
    list_display = ("order", "formatted_product_name", "color_variant", "size_variant", "quantity", "formatted_price", "total_price")
    list_filter = ("order", "product", "color_variant", "size_variant")
    
    def formatted_product_name(self, obj):
        return obj.product.product_name if obj.product else "Deleted Product"
    formatted_product_name.short_description = "Product Name"

    def formatted_price(self, obj):
        return f"₹{obj.price:.2f}"
    formatted_price.short_description = "Price"

    def total_price(self, obj):
        return f"₹{obj.price * obj.quantity:.2f}"
    total_price.short_description = "Total Price"

admin.site.register(OrderItem, OrderItemAdmin)